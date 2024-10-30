"""
七星彩历史数据获取脚本
功能：获取七星彩开奖历史数据并保存到Excel文件
作者：山猫
版本：1.0.0
最后更新：2024-10-30
"""

import os
import sys
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

__version__ = '1.0.0'

class QXCApp(tk.Tk):
    def __init__(self):
        super().__init__()
        
        # 设置图标
        if getattr(sys, 'frozen', False):
            # 打包后的路径
            application_path = sys._MEIPASS
        else:
            # 开发时的路径
            application_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        
        icon_path = os.path.join(application_path, 'resources', 'tc_favicon.ico')
        if os.path.exists(icon_path):
            if sys.platform == 'darwin':
                # macOS 需要特殊处理
                pass  # tkinter on macOS handles icons differently
            else:
                self.iconbitmap(icon_path)
        
        # 初始化UI
        self.setup_ui()
        
        # 设置数据相关
        self.setup_data()
        
        # 首次运行检查
        self.first_run_check()

    def setup_ui(self):
        """设置UI界面"""
        # 设置窗口标题
        self.title("七星彩数据更新工具")
        
        # 设置窗口大小和位置
        self.center_window(400, 200)
        
        # 创建主框架
        main_frame = ttk.Frame(self, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 创建更新按钮
        self.update_button = ttk.Button(
            main_frame, 
            text="更新数据", 
            command=self.update_data
        )
        self.update_button.grid(row=0, column=0, pady=20)
        
        # 创建状态标签
        self.status_label = ttk.Label(
            main_frame, 
            text="准备就绪"
        )
        self.status_label.grid(row=1, column=0, pady=10)
        
        # 创建菜单栏
        self.create_menubar()

    def setup_data(self):
        """设置数据相关的配置"""
        # 设置API URL
        self.url_template = "https://webapi.sporttery.cn/gateway/lottery/getHistoryPageListV1.qry?gameNo=04&provinceId=0&pageSize=30&isVerify=1&pageNo={}"
        
        # 设置请求头
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36'
        }
        
        # 设置数据目录和文件
        self.data_dir = self.get_data_directory()
        self.output_file = os.path.join(self.data_dir, "qxc_history_data_full.xlsx")
        
        # 确保数据目录存在
        os.makedirs(self.data_dir, exist_ok=True)

    def get_data_directory(self):
        """获取数据存储目录"""
        if sys.platform == 'darwin':  # macOS
            docs_dir = os.path.expanduser('~/Documents')
        else:  # Windows
            docs_dir = os.path.expanduser('~/Documents')
        
        app_data_dir = os.path.join(docs_dir, '七星彩数据')
        os.makedirs(app_data_dir, exist_ok=True)
        return app_data_dir

    def center_window(self, width, height):
        """将窗口居中显示"""
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = max(0, (screen_width - width) // 2)
        y = max(0, (screen_height - height) // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')
        self.resizable(False, False)  # 禁止调整窗口大小

    def create_menubar(self):
        """创建菜单栏"""
        menubar = tk.Menu(self)
        
        # 文件菜单
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="更新数据", command=self.update_data)
        file_menu.add_separator()
        file_menu.add_command(label="打开数据文件", command=self.open_data_file)
        file_menu.add_command(label="打开数据文件夹", command=self.open_data_folder)
        menubar.add_cascade(label="文件", menu=file_menu)
        
        # 帮助菜单
        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="关于", command=self.show_about)
        menubar.add_cascade(label="帮助", menu=help_menu)
        
        self.config(menu=menubar)

    def get_latest_draw(self):
        """获取最新一期开奖信息"""
        try:
            response = requests.get(self.url_template.format(1), headers=self.headers, timeout=10)
            response.raise_for_status()
            data = response.json()
            latest_draw = data.get('value', {}).get('list', [])
            if not latest_draw:
                raise ValueError("未获取到开奖数据")
            return latest_draw[0].get('lotteryDrawNum')
        except requests.RequestException as e:
            self.show_error(f"网络请求失败: {str(e)}")
        except (ValueError, KeyError) as e:
            self.show_error(f"数据格式错误: {str(e)}")
        except Exception as e:
            self.show_error(f"获取最新开奖信息失败: {str(e)}")
        return None

    def read_existing_data(self):
        """读取现有数据"""
        try:
            df = pd.read_excel(self.output_file)
            return df, df['期号'].iloc[0]
        except (FileNotFoundError, pd.errors.EmptyDataError):
            return pd.DataFrame(columns=["开奖日期", "期号", "号码1", "号码2", "号码3", "号码4", "号码5", "号码6", "号码7"]), "0"
        except Exception as e:
            self.show_error(f"读取数据时出错: {str(e)}")
            return None, None

    def fetch_new_data(self, start_draw, latest_draw):
        """抓取新的开奖数据"""
        new_rows = []
        page_no = 1
        max_pages = 100  # 添加最大页数限制
        
        while page_no <= max_pages:
            self.status_label.config(text=f"正在抓取第 {page_no} 页的数据...")
            self.update()
            
            try:
                response = requests.get(self.url_template.format(page_no), headers=self.headers, timeout=10)
                response.raise_for_status()
                page_data = response.json().get('value', {}).get('list', [])
                
                if not page_data:
                    break
                    
                for entry in page_data:
                    draw_number = entry.get('lotteryDrawNum', '')
                    if not draw_number:  # 检查期号是否为空
                        continue
                        
                    if int(draw_number) <= int(start_draw):
                        return new_rows
                        
                    draw_date = entry.get('lotteryDrawTime', '')
                    result = entry.get('lotteryDrawResult', '').split(' ')
                    if len(result) != 7:  # 检查号码数量
                        continue
                        
                    new_rows.append([draw_date, draw_number] + result)
                    
                page_no += 1
                
            except Exception as e:
                self.show_error(f"抓取第 {page_no} 页数据时出错: {str(e)}")
                return new_rows
        
        return new_rows

    def update_data(self):
        """更新数据"""
        self.update_button.config(state='disabled')
        self.status_label.config(text="正在检查更新...")
        self.update()
        
        try:
            # 读取现有数据
            existing_df, last_draw = self.read_existing_data()
            if existing_df is None:
                return
                
            # 获取最新一期开奖号码
            latest_draw = self.get_latest_draw()
            if latest_draw is None:
                return
                
            if int(latest_draw) > int(last_draw):
                self.status_label.config(text="发现新数据，开始抓取...")
                self.update()
                
                new_rows = self.fetch_new_data(last_draw, latest_draw)
                
                if new_rows:
                    # 创建新数据的DataFrame
                    new_df = pd.DataFrame(new_rows, columns=["开奖日期", "期号", "号码1", "号码2", "号码3", "号码4", "号码5", "号码6", "号码7"])
                    
                    # 合并新旧数据
                    df = pd.concat([new_df, existing_df], ignore_index=True)
                    
                    # 保存为Excel文件
                    df.to_excel(self.output_file, index=False)
                    
                    # 调整格式
                    wb = load_workbook(self.output_file)
                    ws = wb.active
                    
                    for row in ws.iter_rows():
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    ws.column_dimensions['A'].width = 11
                    wb.save(self.output_file)
                    
                    self.show_success(f"成功添加了 {len(new_rows)} 期新数据！")
                else:
                    self.show_info("没有找到新数据。")
            else:
                self.show_info("数据已是最新，无需更新。")
                
        except Exception as e:
            self.show_error(f"更新数据时出错: {str(e)}")
            
        finally:
            self.update_button.config(state='normal')
            self.status_label.config(text="准备就绪")

    def open_data_file(self):
        """打开数据文件"""
        if sys.platform == 'darwin':  # macOS
            os.system(f'open "{self.output_file}"')
        else:  # Windows
            os.startfile(self.output_file)

    def open_data_folder(self):
        """打开数据文件所在文件夹"""
        if sys.platform == 'darwin':  # macOS
            os.system(f'open -R "{self.output_file}"')
        else:  # Windows
            os.startfile(os.path.dirname(self.output_file))

    def show_about(self):
        """显示关于对话框"""
        about_window = tk.Toplevel(self)
        about_window.title("关于")
        about_window.geometry("300x200")
        
        # 居中显示
        about_window.transient(self)
        about_window.grab_set()
        
        tk.Label(about_window, text="七星彩数据更新工具", font=('', 16, 'bold')).pack(pady=20)
        tk.Label(about_window, text=f"版本 {__version__}").pack()
        tk.Label(about_window, text="© 2024").pack()

    def first_run_check(self):
        """首次运行检查"""
        if not os.path.exists(self.output_file):
            self.show_info(
                "欢迎使用七星彩数据更新工具！\n\n"
                "这似乎是你第一次运行本程序。\n"
                f"数据文件将保存在：\n{self.data_dir}"
            )

    def show_error(self, message):
        """显示错误消息"""
        messagebox.showerror("错误", message)
    
    def show_info(self, message):
        """显示信息消息"""
        messagebox.showinfo("提示", message)
    
    def show_success(self, message):
        """显示成功消息"""
        messagebox.showinfo("成功", message)

def main():
    app = QXCApp()
    app.mainloop()

if __name__ == "__main__":
    main()
